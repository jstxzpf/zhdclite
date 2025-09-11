import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from src.utils import sanitize_filename

class ExcelOperations:
    @staticmethod
    def save_to_excel(df, filename, sheet_name):
        """保存数据到Excel并应用专业格式"""
        # 确保第一列为字符串类型
        if not df.empty:
            df.iloc[:, 0] = df.iloc[:, 0].astype(str)

        # 使用openpyxl直接操作以获得更好的格式控制
        try:
            # 加载现有工作簿
            workbook = openpyxl.load_workbook(filename)

            # 如果工作表已存在，删除它
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]

            # 创建新工作表
            worksheet = workbook.create_sheet(sheet_name)

            # 写入数据
            for r in dataframe_to_rows(df, index=False, header=True):
                worksheet.append(r)

            # 应用格式设置
            ExcelOperations._apply_excel_formatting(worksheet, df)

            # 保存工作簿
            workbook.save(filename)
            workbook.close()

        except Exception:
            # 如果出错，回退到简单方法
            with pd.ExcelWriter(filename, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    @staticmethod
    def _apply_excel_formatting(worksheet, df):
        """应用Excel格式设置"""
        if worksheet.max_row <= 1:  # 只有表头或无数据
            return

        # 定义样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        data_font = Font(name='微软雅黑', size=10)
        data_alignment = Alignment(horizontal='center', vertical='center')

        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 应用表头格式
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 应用数据格式
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

        # 自动调整列宽
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # 设置列宽，最小8，最大50
            adjusted_width = min(max(max_length + 2, 8), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # 设置行高
        for row in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 20

    def _save_df_to_excel(self, df, file_path, sheet_name):
        """
        通用的DataFrame到Excel保存函数，包含格式化。

        Args:
            df (pd.DataFrame): 要保存的数据框
            file_path (str): 完整的文件保存路径
            sheet_name (str): 工作表名称
        """
        # 确保第一列为字符串类型
        if not df.empty and df.shape[1] > 0:
            df.iloc[:, 0] = df.iloc[:, 0].astype(str)

        # 创建工作簿并保存数据
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(r)

        # 应用格式设置
        self._apply_excel_formatting(worksheet, df)

        # 保存工作簿
        workbook.save(file_path)
        workbook.close()



    def save_summary_table(self, df, year, period, category):
        """
        保存汇总表数据到Excel文件

        Args:
            df: 数据框
            year: 年份
            period: 期别 (上半年/下半年)
            category: 类别 (全部/城镇点/农村点)

        Returns:
            str: 生成的文件路径
        """
        try:
            # 构建文件名
            filename = f"{year}年_{period}_{category}_汇总表.xlsx"
            filename = sanitize_filename(filename)

            # 使用应用程序的uploads目录
            output_dir = os.path.join(os.getcwd(), 'uploads')
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, mode=0o777, exist_ok=True)
            file_path = os.path.join(output_dir, filename)

            self._save_df_to_excel(df, file_path, '汇总表')

            return file_path
        except Exception as e:
            raise Exception(f"保存汇总表失败: {str(e)}")

    @staticmethod
    def read_excel(file_path):
        # 先以字符串类型读取，避免数字自动转换为浮点数
        df = pd.read_excel(file_path, dtype=str)

        # 清理可能的 .0 后缀（针对原本是整数的数字列）
        for col in df.columns:
            if df[col].dtype == 'object':  # 字符串列
                # 移除末尾的 .0（如 "3212830010021154.0" -> "3212830010021154"）
                df[col] = df[col].astype(str).str.replace(r'\.0$', '', regex=True)
                # 处理 NaN 值，转换为空字符串
                df[col] = df[col].replace('nan', '')

        return df