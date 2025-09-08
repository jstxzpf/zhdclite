#!/usr/bin/env python3
"""
电子台账Excel文件生成器
专门用于生成格式化的电子台账Excel文件
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
from src.utils import sanitize_filename


class ElectronicLedgerExcel:
    """电子台账Excel文件生成器"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
    def save_electronic_ledger(self, summary_df, detail_df, consumption_df, town, month, year=None):
        """
        保存电子台账到Excel文件
        
        Args:
            summary_df: 汇总表DataFrame
            detail_df: 分户详细账DataFrame
            consumption_df: 分户消费结构DataFrame
            town: 乡镇名称
            month: 月份 (格式: "01", "02", ...)
            year: 年份 (格式: "2024", "2025", ...), 如果为None则使用当前年份
            
        Returns:
            str: 生成的文件路径
        """
        try:
            # 如果没有提供年份，使用当前年份
            if year is None:
                from datetime import datetime
                year = str(datetime.now().year)
            
            # 构建文件名格式：YYYY-MM_乡镇名称_电子台帐.xlsx
            filename = f"{year}-{month.zfill(2)}_{town}_电子台帐.xlsx"
            filename = sanitize_filename(filename)
            
            # 使用应用程序的uploads目录
            output_dir = os.path.join(os.getcwd(), 'uploads')
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, mode=0o777, exist_ok=True)
            file_path = os.path.join(output_dir, filename)
            
            # 删除已存在的文件
            if os.path.exists(file_path):
                os.remove(file_path)
            
            self.logger.info(f"开始生成电子台账Excel文件: {file_path}")
            
            # 创建工作簿
            workbook = openpyxl.Workbook()
            
            # 删除默认工作表
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            
            # 创建三个工作表
            self._create_summary_sheet(workbook, summary_df, "汇总表")
            self._create_detail_sheet(workbook, detail_df, "分户详细账")
            self._create_consumption_sheet(workbook, consumption_df, "分户消费结构")
            
            # 保存文件
            workbook.save(file_path)
            workbook.close()
            
            self.logger.info(f"电子台账Excel文件生成成功: {file_path}")
            return file_path
            
        except Exception as e:
            self.logger.error(f"保存电子台账失败: {str(e)}")
            raise Exception(f"保存电子台账失败: {str(e)}")

    def save_electronic_ledger_to_dir(self, summary_df, detail_df, consumption_df, town, month, year, output_dir):
        """
        保存电子台账到指定目录

        Args:
            summary_df: 汇总表DataFrame
            detail_df: 分户详细账DataFrame
            consumption_df: 分户消费结构DataFrame
            town: 乡镇名称
            month: 月份 (格式: "01", "02", ...)
            year: 年份 (格式: "2024", "2025", ...)
            output_dir: 输出目录路径

        Returns:
            str: 生成的文件路径
        """
        try:
            # 构建文件名格式：YYYY-MM_乡镇名称_电子台帐.xlsx
            filename = f"{year}-{month.zfill(2)}_{town}_电子台帐.xlsx"
            filename = sanitize_filename(filename)

            # 确保输出目录存在
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, mode=0o777, exist_ok=True)

            file_path = os.path.join(output_dir, filename)

            # 删除已存在的文件
            if os.path.exists(file_path):
                os.remove(file_path)

            self.logger.info(f"开始生成电子台账Excel文件: {file_path}")

            # 创建工作簿
            workbook = openpyxl.Workbook()

            # 删除默认工作表
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])

            # 创建三个工作表
            self._create_summary_sheet(workbook, summary_df, "汇总表")
            self._create_detail_sheet(workbook, detail_df, "分户详细账")
            self._create_consumption_sheet(workbook, consumption_df, "分户消费结构")

            # 保存文件
            workbook.save(file_path)
            workbook.close()

            self.logger.info(f"电子台账Excel文件生成成功: {file_path}")
            return file_path

        except Exception as e:
            self.logger.error(f"保存电子台账到指定目录失败: {str(e)}")
            raise Exception(f"保存电子台账到指定目录失败: {str(e)}")
    
    def _create_summary_sheet(self, workbook, df, sheet_name):
        """创建汇总表工作表"""
        worksheet = workbook.create_sheet(sheet_name)
        
        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(r)
        
        # 应用格式设置
        self._apply_summary_formatting(worksheet, df)
        
    def _create_detail_sheet(self, workbook, df, sheet_name):
        """创建分户详细账工作表"""
        worksheet = workbook.create_sheet(sheet_name)
        
        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(r)
        
        # 应用格式设置
        self._apply_detail_formatting(worksheet, df)
        
    def _create_consumption_sheet(self, workbook, df, sheet_name):
        """创建分户消费结构工作表"""
        worksheet = workbook.create_sheet(sheet_name)
        
        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(r)
        
        # 应用格式设置
        self._apply_consumption_formatting(worksheet, df)
    
    def _apply_summary_formatting(self, worksheet, df):
        """应用汇总表格式设置"""
        if df.empty:
            return
            
        # 设置列宽
        column_widths = {
            'A': 15,  # 户代码
            'B': 12,  # 户主姓名
            'C': 12,  # 收入
            'D': 12,  # 支出
            'E': 10,  # 记账笔数
            'F': 12   # 漏记账天数
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 应用通用格式
        self._apply_common_formatting(worksheet, len(df) + 1)
    
    def _apply_detail_formatting(self, worksheet, df):
        """应用分户详细账格式设置"""
        if df.empty:
            return
            
        # 设置列宽
        column_widths = {
            'A': 15,  # 户代码
            'B': 12,  # 户主姓名
            'C': 10,  # 编码
            'D': 10,  # 数量
            'E': 12,  # 金额
            'F': 12,  # 日期
            'G': 10,  # 收支类型
            'H': 10,  # ID
            'I': 20,  # 类型名称
            'J': 12   # 单位名称
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 应用通用格式
        self._apply_common_formatting(worksheet, len(df) + 1)
    
    def _apply_consumption_formatting(self, worksheet, df):
        """应用分户消费结构格式设置"""
        if df.empty:
            return
            
        # 设置列宽
        column_widths = {
            'A': 15,  # 户代码
            'B': 12,  # 户主姓名
            'C': 10,  # 编码
            'D': 25,  # 帐目指标名称
            'E': 12,  # 总金额
            'F': 10   # 记账笔数
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 应用通用格式
        self._apply_common_formatting(worksheet, len(df) + 1)
    
    def _apply_common_formatting(self, worksheet, max_row):
        """应用通用格式设置"""
        # 字体设置
        font = Font(name='宋体', size=10)
        header_font = Font(name='宋体', size=10, bold=True)
        
        # 对齐设置
        alignment = Alignment(horizontal='center', vertical='center')
        
        # 边框设置
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 表头背景色
        header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        
        # 应用格式到所有单元格
        for row in range(1, max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                
                # 设置字体
                if row == 1:  # 表头
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = font
                
                # 设置对齐和边框
                cell.alignment = alignment
                cell.border = thin_border
        
        # 冻结首行
        worksheet.freeze_panes = 'A2'
